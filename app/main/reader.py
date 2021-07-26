import re, os, unicodedata
from collections import defaultdict
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
import nltk
import chardet

from .util import is_supported_file, get_dataset_info

# Split out the levels given a file path relative to a dataset's home directoy
def parse_levels(filename, home):
    relpath = str(Path(filename).absolute().relative_to(home))
    return [x.replace(' ', '_').lower() for x in relpath.split(os.sep)[:-1]]

# Yield all filenames from path.
def get_files(path):
    for filename in Path(path).glob('**/*'):
        if is_supported_file(filename):
            yield filename

# Yield all filenames (recursively) from path, but only if the levels match level_filters.
def get_files_filtered(path, level_filters):
    # It would be more efficient to filter before getting the list of files...
    # Only matters if we have a ton of files...
    for filename in get_files(path):
        levels = parse_levels(filename, path)
        for lev, filt in zip(levels, level_filters):
            if lev not in filt: break
            else: pass
        else:
            yield filename, levels

def get_encoding(filename):
    with open(filename, 'rb') as f:
        return chardet.detect(f.read())['encoding'] # is this slow?

# @TODO: Weird non-unicode character \xa0? I need to investigate further.    
def cleanup_text(text):
    # If in between words, it should be a space.
    text = re.sub(r'(\w)\xa0(\w)', r'\1 \2', text)

    # Otherwise it should be nothing (I think?)
    text = re.sub('\xa0', '', text)

    return text
    

# just get the text
# @TODO: Check plain text file encoding?
def read_text_file(filename, splitter, start, end):
    # @TODO: For huge files, don't read all at once!
    with open(filename, 'r', encoding=get_encoding(filename)) as f:
        doc = cleanup_text(f.read())

    # @TODO: For huge files, use/find a regex splitter that uses a generator.
    # If both start and end aren't empty, always split using those.
    if start != re.compile('',re.M) and end != re.compile('',re.M):
        print('Using start and end!')
        split1 = start.split(doc)
        split2 = []
        for object in split1:
            split2.append(end.split(object))
        for object in split2:
            for thing in object:
                if thing == '':
                    split2.remove(object)
        for objects in split2:
            del objects[-1]
        newlist = []
        for object in split2:
            newlist = newlist + object
        yield from newlist
    # If splitter exists and one/both of start and end are empty, use splitter
    elif splitter != re.compile('',re.M) and (start == re.compile('', re.M) or end == re.compile('', re.M)):
        print('Using splitter!')
        yield from splitter.split(doc)
    # If none, don't split.
    else:
        print('Using none!')
        yield doc

# filename and sheet_name are only for error messages
def load_xlsx_sheet(filename, sheet_name, ws, text_col_pattern):
    def bad(msg):
        print(f"\tWARNING: ({filename}[{sheet_name}]) -- {msg}.")
    data = ws.values
    try:
        header = next(data)
    except StopIteration:
        bad("Sheet is blank; skipping")
        return
    text_col_indices = [i for i in range(len(header)) if \
                        header[i] is not None and \
                        text_col_pattern.search(header[i])]
    if len(text_col_indices) == 0:
        bad(f"No columns match {text_col_pattern.pattern}; skipping.")
        return
    idx = text_col_indices[-1]
    if len(text_col_indices) > 1:
        bad("Multiple text columns; taking last ({idx})")

    for i, row in enumerate(data):
        if row == header: # extra header line?
            continue
        if not any(row): # blank row
            continue
        text = row[idx]
        if not text: # blank cell
            continue
        yield text

# @TODO: Prompt user for text column pattern
DEFAULT_COLUMN = re.compile(r'^Text')
def read_xlsx_file(filename, text_col_pattern=DEFAULT_COLUMN):
    wb = load_workbook(filename)
    for name in wb.sheetnames:
        ws = wb[name]
        yield from load_xlsx_sheet(filename, name, ws, text_col_pattern)

# Yield all articles from filename, using `splitter` to split the file into articles.
def load_file(filename, splitter, start, end):
    assert is_supported_file(filename)
    if str(filename).endswith('xlsx'):
        yield from read_xlsx_file(filename)
    else:
        yield from read_text_file(filename, splitter, start, end)

# @TODO: Don't assume two line breaks mark a paragraph!  Sometimes not
# true, e.g., Word uses single line break.  I don't really know how to
# fix it in the general case; we'd have to be able to detect
# paragraphs semantically.
DEFAULT_PARA_SPLITTER = re.compile(r'(?:\n){2,}|(?:\r\n){2,}|(?:\r){2,}')
def get_paragraphs(text, fpat=None, para_splitter=DEFAULT_PARA_SPLITTER):
    for para in para_splitter.split(text):
        if not fpat:
            if len(para) > 3: # minimum length (arbitrary?)
                yield para
            else:
                pass
        elif fpat.search(para):
            yield para

def get_naive_sentences(text, fpat=None):
    for para in get_paragraphs(text):
        for sent in nltk.tokenize.sent_tokenize(para):
            if not fpat or fpat.search(sent):
                yield sent

QUOTE_PATTERN = re.compile(r'"([^"]*)"')
QUOTE_MARKER = '"|QUOTE| {}"' # keep the quotes to help sent_tokenize
def get_sentences(text, fpat=None, smart=False):
    if not smart:
        yield from get_naive_sentences(text, fpat)
        return
    
    # Try to handle quotations correctly. We assume a quotation never
    # appears standalone, as its own sentence.
    # 
    # Similar to [this](https://stackoverflow.com/questions/32003294/sentence-tokenization-for-texts-that-contains-quotes)
    #
    # The problem is that it is brittle, since there are often
    # mistakes in news articles, e.g., quotes that are not closed. I
    # do not have time to mess with this, so right now the default is
    # just to use nltk.

    for para in get_paragraphs(text):
        quotations = QUOTE_PATTERN.findall(para) # save for later
        para_unquote = para
        for q in quotations:
            # include the last little bit, which may include punctuation to help sent_tokenize.
            marker = QUOTE_MARKER.format(q[-3:])
            para_unquote = para_unquote.replace(f'"{q}"', marker)

        # Now tokenize without them messing everything up. I could do
        # this just by yielding the sentences, but I predict this
        # bespoke approach to handling quotes will occasionally fail
        # in real life, so let's check for errors and fall back to
        # naive nltk when necessary.
        sents = []
        qi = 0
        for sent in nltk.tokenize.sent_tokenize(para_unquote):
            marker_start = QUOTE_MARKER.split()[0]
            while marker_start in sent:
                if qi >= len(quotations):
                    # @DEBUGGING
                    print(f"(WARNING)\t> Too many quotations? ({qi})")
                    print(para)
                    print('---')
                    print(quotations)
                    print('---')
                    print(para_unquote)
                q = quotations[qi]
                marker = QUOTE_MARKER.format(q[-3:])
                sent = sent.replace(marker, f'"{q}"')
                qi += 1
            if not fpat or fpat.search(sent):
                sents.append(sent) # yield sent
        
        if qi < len(quotations):
            print("WARNING: custom quote handling failed; quotes remain!")
            print(quotations[qi:])
            print(para)
            for sent in nltk.tokenize.sent_tokenize(para):
                if not fpat or fpat.search(sent):
                    yield sent
        else:
            yield from sents

def get_fixed_windows(text, fpat=None, csize=2):
    sents = list(get_sentences(text))
    wsize = csize*2 + 1
    for start in range(0, len(sents), wsize):
        end = min(start+wsize, len(sents))
        chunk = ' '.join(sents[start:end])
        if not fpat or fpat.search(chunk):
            yield chunk
            
# Find the pattern, then give X sentences before and after
def get_contexts_no_overlap(text, fpat, csize=2):
    assert fpat # doesn't make sense without a pattern
    sents = list(get_sentences(text))
    i, prev = 0, 0
    while i < len(sents):
        if fpat.search(sents[i]):
            start = max(prev, i-csize)
            end = min(i+csize, len(sents)-1)
            yield ' '.join(sents[start:end+1])
            i = end+1
            prev = end+1
        else:
            i += 1

def get_contexts_with_overlap(text, fpat, csize=2):
    assert fpat # doesn't make sense without a pattern
    sents = list(get_sentences(text))

    for i in range(len(sents)):
        if fpat.search(sents[i]):
            start = max(0, i-csize)
            end = min(i+csize, len(sents)-1)
            win = ' '.join(sents[start:end+1])
            yield win

# Like get_contexts_no_overlap, except if we see the pattern in the
# 'edge' sentences, we should expand the window by min_csize and
# continue until we no longer see the pattern. This mostly means
# moving 'down' the document, but should we take as much text above,
# to make sure we get the leading discussion?
def get_expandable_contexts(text, fpat, min_csize=2):
    raise NotImplementedError()

# fpat is not used
def get_article(text, fpat=None):
    yield 

def get_chunks(article, unit, fpat=None):
    # These all need to be generators!
    parsers = {
        'fixed_windows': get_fixed_windows,
        'paragraphs': get_paragraphs,
        'articles': get_article,
        'context_no_overlap': get_contexts_no_overlap,
        'context_with_overlap': get_contexts_with_overlap,
        'sentences': get_sentences
    }
    assert unit in parsers, f"Bad unit: {unit}"
    yield from parsers[unit](article, fpat)

def load_raw_articles(path, level_names, splitter, start, end):
    articles = {'Filename': [], 'Text': [], 'Article ID': []}
    articles.update({f'_level_{i}': [] for i in range(len(level_names))})

    for filename in get_files(path):
        levels = parse_levels(filename, path)
        for aid, article in enumerate(load_file(filename, splitter, start, end)):
            articles['Filename'].append(filename.relative_to(path))
            articles['Text'].append(article)
            articles['Article ID'].append(aid)

            for i in range(len(level_names)):
                val = levels[i] if i < len(levels) else pd.NA
                articles[f'_level_{i}'].append(val)
    return pd.DataFrame(articles)
            
def load_wrangled(home, level_filters, uoa, fpattern):
    info = get_dataset_info(home)
    splitter = info['article_regex_splitter']
    level_names = info['level_names']
    start = info['start_regex_splitter']
    end = info['start_regex_splitter']

    # I set all these to empty lists instead of defaultdict to make
    # sure we have all the right columns.
    chunks = {'Text': [], 'Article ID': [], 'Filename': []}
    level_placeholders = [f'_level_{i}' for i in range(len(level_names))]
    chunks.update({f'{name}': [] for name in level_placeholders})

    wrangledpath= '.wrangled.pkl'
    cache = os.path.join(home, wrangledpath)
    if os.path.exists(cache):
        all_articles = pd.read_pickle(cache)
    else:
        all_articles = load_raw_articles(home, level_names, splitter, start, end)

    df = all_articles
    for lname, filt in zip(level_placeholders, level_filters):
        # if NA, then it means any?
        df = df[df[lname].isin(filt) | df[lname].isna()]

    for _, row in df.iterrows():
        for chunk in get_chunks(row['Text'], uoa, fpattern):
            for col in df.columns:
                chunks[col].append(row[col])
            chunks['Text'][-1] = chunk
            
    del all_articles['Text']
    return all_articles, pd.DataFrame(chunks)
                
import unittest

class TestChunkers(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls._text = """Here is the first paragraph. Second sentence. And now here is a quote: "I am a quote. Do not question it." Another sentence, with more text.

This is the second paragraph. (I am sure.) But there are only three sentences (this parenthetical comment is included in the current sentence, but the previous parenthetical is it's own sentence)!

And is this the third paragraph? I need more sentences in this one. The original dataset involved the 2020 presidential election. The candidates were Donald Trump and Joseph Biden. Mr. Trump's running mate was Mr. Mike Pence, while Mr. Biden's vice presidential running mate was Mrs. Kamala Harris. And this is the sixth sentence. Today Congress will certify the electoral votes, though some Republican congress members plan to object. I do not know (question) the rules regarding such things.

Well, that's it. And now we have the final paragraph. "The end," said Dr. Utterback."""
        
    def test_fixed_windows(self):
        text = type(self)._text
        windows = list(get_fixed_windows(text))
        self.assertEqual(len(windows), 4)
        self.assertTrue(windows[0].startswith("Here is the first paragraph."))
        self.assertTrue(windows[0].endswith("This is the second paragraph."))        
        self.assertTrue(windows[1].startswith("(I am sure.) But there"))
        self.assertTrue(windows[1].endswith("2020 presidential election."))        
        self.assertTrue(windows[2].startswith("The candidates were Donald Trump"))
        self.assertTrue(windows[2].endswith("the rules regarding such things."))        
        self.assertTrue(windows[3].startswith("Well, that's it. And now we have"))
        self.assertTrue(windows[3].endswith('"The end," said Dr. Utterback.'))        

    def test_paragraphs(self):
        text = type(self)._text
        paras = list(get_paragraphs(text))
        self.assertEqual(len(paras), 4)
        self.assertTrue(paras[0].startswith("Here is the first paragraph."))
        self.assertTrue(paras[0].endswith("Another sentence, with more text."))
        self.assertTrue(paras[1].startswith("This is the second"))
        self.assertTrue(paras[1].endswith("is it's own sentence)!"))
        self.assertTrue(paras[2].startswith("And is this the"))
        self.assertTrue(paras[2].endswith("rules regarding such things."))
        self.assertTrue(paras[3].startswith("Well, that's it. And now"))
        self.assertTrue(paras[3].endswith("Dr. Utterback."))

        paras = list(get_paragraphs(text, re.compile(r'\bquestion\b', re.I)))
        self.assertEqual(len(paras), 2)
        self.assertTrue(paras[0].startswith("Here is the first"))
        self.assertTrue(paras[0].endswith("with more text."))
        self.assertTrue(paras[1].startswith("And is this the third"))
        self.assertTrue(paras[1].endswith("regarding such things."))
                    
                        
    def test_context_no_overlap(self):
        text = type(self)._text
        pat = re.compile(r'\b(?:Trump|Biden)\b', re.I)
        chunks = list(get_contexts_no_overlap(text, pat))
        self.assertEqual(len(chunks), 1)
        self.assertEqual(chunks[0], """I need more sentences in this one. The original dataset involved the 2020 presidential election. The candidates were Donald Trump and Joseph Biden. Mr. Trump's running mate was Mr. Mike Pence, while Mr. Biden's vice presidential running mate was Mrs. Kamala Harris. And this is the sixth sentence.""")

        pat = re.compile(r'\bquestion\b', re.I)
        chunks = list(get_contexts_no_overlap(text, pat))
        self.assertEqual(len(chunks), 2)
        self.assertEqual(chunks[0], """Here is the first paragraph. Second sentence. And now here is a quote: "I am a quote. Do not question it." Another sentence, with more text. This is the second paragraph.""")
        self.assertEqual(chunks[1], """And this is the sixth sentence. Today Congress will certify the electoral votes, though some Republican congress members plan to object. I do not know (question) the rules regarding such things. Well, that's it. And now we have the final paragraph.""")

        pat = re.compile(r'\bparagraph\b', re.I)
        chunks = list(get_contexts_no_overlap(text, pat))
        self.assertEqual(len(chunks), 4)
        self.assertEqual(chunks[0], 'Here is the first paragraph. Second sentence. And now here is a quote: "I am a quote. Do not question it."')
        self.assertEqual(chunks[1], """Another sentence, with more text. This is the second paragraph. (I am sure.) But there are only three sentences (this parenthetical comment is included in the current sentence, but the previous parenthetical is it's own sentence)!""")
        self.assertEqual(chunks[2], "And is this the third paragraph? I need more sentences in this one. The original dataset involved the 2020 presidential election.")
        self.assertEqual(chunks[3], """I do not know (question) the rules regarding such things. Well, that's it. And now we have the final paragraph. "The end," said Dr. Utterback.""")
        
    def test_context_with_overlap(self):
        text = type(self)._text
        pat = re.compile(r'\b(?:Trump|Biden)\b', re.I)
        chunks = list(get_contexts_with_overlap(text, pat))
        self.assertEqual(len(chunks), 2)
        self.assertEqual(chunks[0], """I need more sentences in this one. The original dataset involved the 2020 presidential election. The candidates were Donald Trump and Joseph Biden. Mr. Trump's running mate was Mr. Mike Pence, while Mr. Biden's vice presidential running mate was Mrs. Kamala Harris. And this is the sixth sentence.""")
        self.assertEqual(chunks[1], """The original dataset involved the 2020 presidential election. The candidates were Donald Trump and Joseph Biden. Mr. Trump's running mate was Mr. Mike Pence, while Mr. Biden's vice presidential running mate was Mrs. Kamala Harris. And this is the sixth sentence. Today Congress will certify the electoral votes, though some Republican congress members plan to object.""")

        pat = re.compile(r'\bquestion\b', re.I)
        chunks = list(get_contexts_with_overlap(text, pat))
        self.assertEqual(len(chunks), 2)
        self.assertEqual(chunks[0], """Here is the first paragraph. Second sentence. And now here is a quote: "I am a quote. Do not question it." Another sentence, with more text. This is the second paragraph.""")
        self.assertEqual(chunks[1], """And this is the sixth sentence. Today Congress will certify the electoral votes, though some Republican congress members plan to object. I do not know (question) the rules regarding such things. Well, that's it. And now we have the final paragraph.""")

        pat = re.compile(r'\bparagraph\b', re.I)
        chunks = list(get_contexts_with_overlap(text, pat))
        self.assertEqual(len(chunks), 4)
        self.assertEqual(chunks[0], 'Here is the first paragraph. Second sentence. And now here is a quote: "I am a quote. Do not question it."')
        self.assertEqual(chunks[1], """And now here is a quote: "I am a quote. Do not question it." Another sentence, with more text. This is the second paragraph. (I am sure.) But there are only three sentences (this parenthetical comment is included in the current sentence, but the previous parenthetical is it's own sentence)!""")
        self.assertEqual(chunks[2], "(I am sure.) But there are only three sentences (this parenthetical comment is included in the current sentence, but the previous parenthetical is it's own sentence)! And is this the third paragraph? I need more sentences in this one. The original dataset involved the 2020 presidential election.")
        self.assertEqual(chunks[3], """I do not know (question) the rules regarding such things. Well, that's it. And now we have the final paragraph. "The end," said Dr. Utterback.""")

    
    def test_sentences(self):
        text = type(self)._text

        # without pattern
        sents = list(get_sentences(text))
        self.assertEqual(len(sents), 18)
        self.assertEqual(sents[0], "Here is the first paragraph.")
        self.assertEqual(sents[1], "Second sentence.")
        self.assertEqual(sents[2], 'And now here is a quote: "I am a quote. Do not question it."')
        self.assertEqual(sents[3], "Another sentence, with more text.")

        
        self.assertEqual(sents[4], "This is the second paragraph.")
        self.assertEqual(sents[5], "(I am sure.)")
        self.assertEqual(sents[6], "But there are only three sentences (this parenthetical comment is included in the current sentence, but the previous parenthetical is it's own sentence)!")

        self.assertEqual(sents[7], "And is this the third paragraph?")
        self.assertEqual(sents[8], "I need more sentences in this one.")
        self.assertEqual(sents[9], "The original dataset involved the 2020 presidential election.")
        self.assertEqual(sents[10], "The candidates were Donald Trump and Joseph Biden.")
        self.assertEqual(sents[11], "Mr. Trump's running mate was Mr. Mike Pence, while Mr. Biden's vice presidential running mate was Mrs. Kamala Harris.")
        self.assertEqual(sents[12], "And this is the sixth sentence.")
        self.assertEqual(sents[13], "Today Congress will certify the electoral votes, though some Republican congress members plan to object.")
        self.assertEqual(sents[14], "I do not know (question) the rules regarding such things.")
        self.assertEqual(sents[15], "Well, that's it.")
        self.assertEqual(sents[16], "And now we have the final paragraph.")
        self.assertEqual(sents[17], '"The end," said Dr. Utterback.')

        # with pattern
        sents = list(get_sentences(text, re.compile(r'\b(?:Trump|Biden)\b', re.I)))
        self.assertEqual(len(sents), 2)
        self.assertEqual(sents[0], "The candidates were Donald Trump and Joseph Biden.")
        self.assertEqual(sents[1], "Mr. Trump's running mate was Mr. Mike Pence, while Mr. Biden's vice presidential running mate was Mrs. Kamala Harris.")
        


if __name__ == '__main__':
    unittest.main()                     
